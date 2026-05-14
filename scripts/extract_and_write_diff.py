#!/usr/bin/env python3
"""
差分自動抽出・Excel転記スクリプト
usage:
  python3 extract_and_write_diff.py --project ver1 --variant m02 \
      --old <旧解析結果ディレクトリ> \
      --new <新解析結果ディレクトリ> \
      --xlsx <Excelファイルパス> \
      [--dry-run]

--dry-run を付けると Excel には書き込まず、抽出結果のみ表示する。
"""
import argparse
import csv
import difflib
import re
import warnings
from pathlib import Path

import openpyxl

LINK_TEXT = "過去プロジェクト(ver1/ver2/ver2.1)の 要件ファイルフォーマット変更"
SHEET_NAME = "Output差分"

# ──────────────────────────────────────────
# ユーティリティ
# ──────────────────────────────────────────

def read_csv_rows(path: Path) -> list[dict]:
    with open(path, encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def read_lines(path: Path) -> list[str]:
    with open(path, encoding="utf-8") as f:
        return f.readlines()


def image_list(folder: Path) -> set[str]:
    if not folder.exists():
        return set()
    return {f.name for f in folder.iterdir() if f.is_file()}


def count_same_clock_reorders(old_path: Path, new_path: Path) -> int:
    """start_clock_ms が同一の行が旧→新で順序入れ替わった件数を近似カウントする"""
    old_lines = read_lines(old_path)
    new_lines = read_lines(new_path)
    changed = sum(
        1 for tag, i1, i2, j1, j2
        in difflib.SequenceMatcher(None, old_lines, new_lines).get_opcodes()
        if tag == "replace"
        for _ in range(min(i2 - i1, j2 - j1))
    )
    return changed


# ──────────────────────────────────────────
# 差分抽出ロジック群
# ──────────────────────────────────────────

def extract_input_info(old_dir: Path, new_dir: Path) -> list[dict]:
    """input_info.txt: ツール差分・要件ファイル差分の2件を固定出力"""
    entries = []
    old_path = old_dir / "input_info.txt"
    new_path = new_dir / "input_info.txt"
    if not old_path.exists() or not new_path.exists():
        return entries
    if old_path.read_text(encoding="utf-8") == new_path.read_text(encoding="utf-8"):
        return entries

    old_tool = ""
    new_tool = ""
    old_req = ""
    new_req = ""
    tag_map = {"Tool Tag:": "tool_old", "Tool Branch:": "tool_new",
               "chksimyml:": "req"}
    for line in old_path.read_text(encoding="utf-8").splitlines():
        if line.startswith("Tool Tag:") or line.startswith("Tool Branch:"):
            old_tool = line.split(":", 1)[1].strip()
        if line.startswith("chksimyml:"):
            old_req = line.split(":", 1)[1].strip()
    for line in new_path.read_text(encoding="utf-8").splitlines():
        if line.startswith("Tool Tag:") or line.startswith("Tool Branch:"):
            new_tool = line.split(":", 1)[1].strip()
        if line.startswith("chksimyml:"):
            new_req = line.split(":", 1)[1].strip()

    entries.append({
        "folder": "-", "file": "input_info.txt",
        "diff_type": "入力ファイルの差分",
        "old_val": "-", "new_val": "要件チェックツール差分",
        "cause": "ver3ツールを用いたことによる差分",
    })
    entries.append({
        "folder": "-", "file": "input_info.txt",
        "diff_type": "入力ファイルの差分",
        "old_val": "-", "new_val": "要件ファイル差分",
        "cause": "ver3フォーマットを使用した要件ファイルに置き換わったことによる差分",
    })
    return entries


def extract_image_diffs(old_dir: Path, new_dir: Path) -> list[dict]:
    """グラフ画像フォルダ: ファイル消失・リネーム（画像差分）を抽出する"""
    graph_dirs = [
        "processing_load_duration_graph",
        "Sequence_duration_graph",
        "SWC_budget_duration_graph",
        "WakeupInterval_graph",
    ]
    entries = []
    for gdir in graph_dirs:
        for sub in ("PASS", "FAIL"):
            old_imgs = image_list(old_dir / gdir / sub)
            new_imgs = image_list(new_dir / gdir / sub)
            deleted = sorted(old_imgs - new_imgs)
            added = sorted(new_imgs - old_imgs)

            # 削除のみ → ファイル消失
            for img in deleted:
                # 対応する追加候補がなければ「消失」
                matched = False
                for a in added:
                    if difflib.SequenceMatcher(None, img, a).ratio() > 0.7:
                        matched = True
                        break
                if not matched:
                    entries.append({
                        "folder": gdir, "file": img,
                        "diff_type": "ファイル消失",
                        "old_val": f"{img} が存在する",
                        "new_val": f"{img} が存在しない",
                        "cause": None,
                    })

            # 高類似ペア → 画像差分（リネーム）
            for d_img in deleted:
                for a_img in added:
                    ratio = difflib.SequenceMatcher(None, d_img, a_img).ratio()
                    if ratio > 0.7:
                        # タイトルの変化を簡潔に記述
                        old_stem = Path(d_img).stem
                        new_stem = Path(a_img).stem
                        entries.append({
                            "folder": gdir, "file": d_img,
                            "diff_type": "画像差分",
                            "old_val": f"タイトル{old_stem}",
                            "new_val": f"タイトル{new_stem}に変化",
                            "cause": None,
                        })
    return entries


def extract_input_data_csv(old_dir: Path, new_dir: Path, filename: str) -> list[dict]:
    """input_data_ba/igr.csv: pf_1ms消失・順序入れ替わりを抽出する"""
    entries = []
    old_path = old_dir / filename
    new_path = new_dir / filename
    if not old_path.exists() or not new_path.exists():
        return entries
    if old_path.read_bytes() == new_path.read_bytes():
        return entries

    # pf_1ms_base / pf_1ms_mid の行数を確認
    old_lines = read_lines(old_path)
    new_lines = read_lines(new_path)

    for node in ("pf_1ms_base", "pf_1ms_mid"):
        old_count = sum(1 for l in old_lines if node in l)
        new_count = sum(1 for l in new_lines if node in l)
        if old_count > 0 and new_count == 0:
            entries.append({
                "folder": "-", "file": filename,
                "diff_type": f"{node}行の消失",
                "old_val": "-", "new_val": f"node={node} の行が消失",
                "cause": None,
            })

    # 順序入れ替わり件数（pf_1ms を除き、同一 start_clock_ms の行が変わった数を近似）
    # node 列が変化している行数をカウントする
    import csv as _csv
    def _count_reorders(path_a, path_b):
        with open(path_a, encoding="utf-8", newline="") as f:
            rows_a = list(_csv.DictReader(f))
        with open(path_b, encoding="utf-8", newline="") as f:
            rows_b = list(_csv.DictReader(f))
        # pf_1ms を除外して行数を揃える
        rows_a = [r for r in rows_a if "pf_1ms" not in r.get("node", "")]
        rows_b = [r for r in rows_b if "pf_1ms" not in r.get("node", "")]
        count = 0
        for ra, rb in zip(rows_a, rows_b):
            if ra.get("start_clock_ms") == rb.get("start_clock_ms") and ra.get("node") != rb.get("node"):
                count += 1
        return count
    reorder_count = _count_reorders(old_path, new_path)
    if reorder_count > 0:
        entries.append({
            "folder": "-", "file": filename,
            "diff_type": "Nodeの順序入れ替わり",
            "old_val": "-",
            "new_val": f"start_clock_msが同じnodeに対して，順序が入れ替わるケースあり({reorder_count}件)",
            "cause": None,
        })
    return entries


def extract_processing_time_result(old_dir: Path, new_dir: Path) -> list[dict]:
    """processing_time_result.csv: PF_window追加を抽出する"""
    entries = []
    fname = "processing_time_result.csv"
    old_path = old_dir / fname
    new_path = new_dir / fname
    if not old_path.exists() or not new_path.exists():
        return entries

    old_lines = set(read_lines(old_path))
    new_lines = set(read_lines(new_path))
    added = new_lines - old_lines
    pf_window_added = any("PF_window" in l for l in added)

    if pf_window_added:
        entries.append({
            "folder": "-", "file": fname,
            "diff_type": "TaskIDの追加",
            "old_val": "-", "new_val": "PF_windowが追加",
            "cause": "infsimyml のフォーマット変更(1ｰ③)に伴う変化点",
        })
    return entries


def _detect_name_changes(old_rows: list[dict], new_rows: list[dict], fields: list[str]) -> list[tuple]:
    """旧→新で変化した (フィールド名, 旧値, 新値) のペアをユニーク抽出する"""
    changes = set()
    old_map = {r.get(fields[0], ""): r for r in old_rows if r.get(fields[0])}
    for new_row in new_rows:
        key = new_row.get(fields[0], "")
        if key in old_map:
            old_row = old_map[key]
            for f in fields:
                ov = old_row.get(f, "")
                nv = new_row.get(f, "")
                if ov != nv and ov and nv:
                    changes.add((f, ov, nv))
    return sorted(changes)


def extract_requirements_csv(old_dir: Path, new_dir: Path, filename: str) -> list[dict]:
    """before/after_requirements.csv: 各種名称変化・キー追加を抽出する"""
    entries = []
    old_path = old_dir / "temp" / filename
    new_path = new_dir / "temp" / filename
    if not old_path.exists() or not new_path.exists():
        return entries
    if old_path.read_bytes() == new_path.read_bytes():
        return entries

    old_rows = read_csv_rows(old_path)
    new_rows = read_csv_rows(new_path)

    old_fields = set(old_rows[0].keys()) if old_rows else set()
    new_fields = set(new_rows[0].keys()) if new_rows else set()
    added_fields = new_fields - old_fields
    if added_fields:
        # Fsync用Targetキー追加など
        entries.append({
            "folder": "tmp", "file": Path(filename).stem,
            "diff_type": "キーの追加",
            "old_val": "-",
            "new_val": f"RequirementType: Fsync に対して Targetキーが追加",
            "cause": "chksimyml のフォーマット変更(2ｰ②)に伴う変化点",
        })

    # Sequence列の変化（Node名統一）
    seq_changes = set()
    for old_r, new_r in zip(old_rows, new_rows):
        for f in ("Sequence", "Sender", "Receiver", "FirstTask", "SecondTask"):
            ov = old_r.get(f, "")
            nv = new_r.get(f, "")
            if ov != nv and (ov or nv):
                seq_changes.add((ov or "-", nv or "-"))
    if seq_changes:
        entries.append({
            "folder": "tmp", "file": Path(filename).stem,
            "diff_type": "Sequence/Sender/Receiver/FirstTask/SecondTaskの変化",
            "old_val": "chksimymlに記載のSequence/Sender/Receiver/FirstTask/SecondTask名",
            "new_val": "SWC名/Node名に統一",
            "cause": "chksimyml のフォーマット変更(2ｰ①)に伴う変化点",
        })

    # RequirementId / RequirementOwner の変化をユニーク抽出
    id_changes: set[tuple] = set()
    owner_changes: set[tuple] = set()
    for old_r, new_r in zip(old_rows, new_rows):
        oid = old_r.get("RequirementId", "") or old_r.get("RequirementID", "")
        nid = new_r.get("RequirementId", "") or new_r.get("RequirementID", "")
        if oid != nid and oid and nid:
            id_changes.add((oid, nid))
        oown = old_r.get("RequirementOwner", "")
        nown = new_r.get("RequirementOwner", "")
        if oown != nown and oown and nown:
            owner_changes.add((oown, nown))

    for old_val, new_val in sorted(id_changes):
        entries.append({
            "folder": "tmp", "file": Path(filename).stem,
            "diff_type": "RequirementIdの変化",
            "old_val": old_val, "new_val": f"{new_val}に変化",
            "cause": "chksimyml のフォーマット変更(2ｰ③)に伴う変化点",
        })
    for old_val, new_val in sorted(owner_changes):
        entries.append({
            "folder": "tmp", "file": Path(filename).stem,
            "diff_type": "RequirementOwnerの変化",
            "old_val": old_val, "new_val": new_val,
            "cause": None,
        })
    return entries


def extract_cpuload_requirements(old_dir: Path, new_dir: Path, filename: str) -> list[dict]:
    """before/after_cpuload_requirements.csv: NodeName変化・PF_window追加を抽出する"""
    entries = []
    old_path = old_dir / "temp" / filename
    new_path = new_dir / "temp" / filename
    if not old_path.exists() or not new_path.exists():
        return entries
    if old_path.read_bytes() == new_path.read_bytes():
        return entries

    old_rows = read_csv_rows(old_path)
    new_rows = read_csv_rows(new_path)

    old_names = {r.get("DisplayName", r.get("NodeName", "")) for r in old_rows}
    new_names = {r.get("DisplayName", r.get("NodeName", "")) for r in new_rows}
    added_names = sorted(new_names - old_names)

    pf_window_added = [n for n in added_names if "PF_window" in n]
    other_added = [n for n in added_names if "PF_window" not in n]

    if pf_window_added:
        entries.append({
            "folder": "tmp", "file": Path(filename).stem,
            "diff_type": "TaskIDの追加",
            "old_val": "-", "new_val": "PF_window が追加",
            "cause": "infsimyml のフォーマット変更(1ｰ③)に伴う変化点",
        })

    # NodeName の変化（旧→新で対応する行の DisplayName が変わった）
    # TaskID をキーに NodeName の変化を抽出する
    name_changes: set[tuple] = set()
    old_by_task = {r.get("TaskID", ""): r for r in old_rows if r.get("TaskID")}
    for r in new_rows:
        tid = r.get("TaskID", "")
        if tid in old_by_task:
            on = old_by_task[tid].get("NodeName", "")
            nn = r.get("NodeName", "")
            if on != nn and on and nn:
                name_changes.add((on, nn))

    for old_val, new_val in sorted(name_changes):
        entries.append({
            "folder": "tmp", "file": Path(filename).stem,
            "diff_type": "NodeNameの変化",
            "old_val": old_val, "new_val": new_val,
            "cause": "infsimyml のフォーマット変更(1ｰ①)に伴う変化点",
        })
    return entries


def extract_budget_csv(old_dir: Path, new_dir: Path, filename: str) -> list[dict]:
    """before/after_budget.csv: TaskList変化を抽出する"""
    entries = []
    old_path = old_dir / "temp" / filename
    new_path = new_dir / "temp" / filename
    if not old_path.exists() or not new_path.exists():
        return entries
    if old_path.read_bytes() == new_path.read_bytes():
        return entries

    entries.append({
        "folder": "tmp", "file": Path(filename).stem,
        "diff_type": "TaskListの変化",
        "old_val": "budget.yamlに記載のTaskList",
        "new_val": "TaskListがNode名で統一",
        "cause": "budget.yamlのフォーマット変更(4ｰ②)に伴う変化点",
    })
    return entries


def extract_tsync_csv(old_dir: Path, new_dir: Path, filename: str) -> list[dict]:
    """before/after_csv_data_tsync_PlusBA.csv: pf_1ms消失・mid時刻値変化を抽出する"""
    entries = []
    old_path = old_dir / "temp" / filename
    new_path = new_dir / "temp" / filename
    if not old_path.exists() or not new_path.exists():
        return entries
    if old_path.read_bytes() == new_path.read_bytes():
        return entries

    old_lines = read_lines(old_path)
    new_lines = read_lines(new_path)

    for node in ("pf_1ms_base", "pf_1ms_mid"):
        old_count = sum(1 for l in old_lines if node in l)
        new_count = sum(1 for l in new_lines if node in l)
        if old_count > 0 and new_count == 0:
            entries.append({
                "folder": "tmp", "file": Path(filename).stem,
                "diff_type": f"{node}行の消失",
                "old_val": "-", "new_val": f"node={node} の行が消失",
                "cause": None,
            })

    # mid行の時刻値変化（pf_1ms以外で変化あり）
    non_pf_old = [l for l in old_lines if "pf_1ms" not in l]
    non_pf_new = [l for l in new_lines if "pf_1ms" not in l]
    changed = any(
        tag == "replace"
        for tag, *_ in difflib.SequenceMatcher(None, non_pf_old, non_pf_new).get_opcodes()
    )
    if changed:
        entries.append({
            "folder": "tmp", "file": Path(filename).stem,
            "diff_type": "mid行の時刻値変化",
            "old_val": "-", "new_val": "mid 行の tsync 補正時刻値が変化",
            "cause": "infsimyml のフォーマット変更(1ｰ③)に伴う変化点",
        })
    return entries


def extract_schedule_result(old_dir: Path, new_dir: Path) -> list[dict]:
    """schedule_result.csv: id/name変化・etc変化を抽出する"""
    entries = []
    fname = "schedule_result.csv"
    old_path = old_dir / fname
    new_path = new_dir / fname
    if not old_path.exists() or not new_path.exists():
        return entries
    if old_path.read_bytes() == new_path.read_bytes():
        return entries

    old_rows = read_csv_rows(old_path)
    new_rows = read_csv_rows(new_path)

    # id / name の変化をユニーク抽出
    id_name_changes: set[tuple] = set()
    etc_changes = False
    for old_r, new_r in zip(old_rows, new_rows):
        for f in ("id", "name"):
            ov = old_r.get(f, "")
            nv = new_r.get(f, "")
            if ov != nv and ov and nv:
                id_name_changes.add((ov, nv))
        ov_etc = old_r.get("etc", "")
        nv_etc = new_r.get("etc", "")
        if ov_etc != nv_etc:
            etc_changes = True

    if id_name_changes:
        # 「VidDraw」「VidDraw-01」のような同一根拠の変化はまとめる
        # 先頭エントリに推定原因を付ける
        # id/name 両方変化したペアは重複除去（値が同じなら1件に統合）
        seen_pairs: set[tuple] = set()
        deduped: list[tuple] = []
        for old_val, new_val in sorted(id_name_changes):
            # 親名称（末尾の -NN 等を除去）が同じならまとめる
            old_base = re.sub(r'-\d+$', '', old_val)
            new_base = re.sub(r'-\d+$', '', new_val)
            key = (old_base, new_base)
            if key not in seen_pairs:
                seen_pairs.add(key)
                deduped.append((old_val, new_val))
        first = True
        for old_val, new_val in deduped:
            entries.append({
                "folder": "-", "file": fname,
                "diff_type": "id/nameの変化",
                "old_val": old_val, "new_val": f"{new_val}",
                "cause": "chksimyml のフォーマット変更(2ｰ③)に伴う変化点" if first else None,
            })
            first = False

    if etc_changes:
        entries.append({
            "folder": "-", "file": fname,
            "diff_type": "etcの変化",
            "old_val": "chksimymlに記載のNode名称",
            "new_val": "SWC名/Node名に統一",
            "cause": "chksimyml のフォーマット変更(2ｰ①)に伴う変化点",
        })
    return entries


def extract_schedule_fail_list(old_dir: Path, new_dir: Path) -> list[dict]:
    """schedule_result_fail_list.csv: 列ヘッダ変化を抽出する"""
    entries = []
    fname = "schedule_result_fail_list.csv"
    old_path = old_dir / fname
    new_path = new_dir / fname
    if not old_path.exists() or not new_path.exists():
        return entries
    if old_path.read_bytes() == new_path.read_bytes():
        return entries

    # ヘッダ（1行目）またはサブヘッダ（2行目）の変化を検出する
    old_lines = open(old_path, encoding="utf-8").readlines()
    new_lines = open(new_path, encoding="utf-8").readlines()
    header_changed = (
        (old_lines[0].rstrip() != new_lines[0].rstrip()) if old_lines and new_lines else False
    )
    subheader_changed = (
        len(old_lines) > 1 and len(new_lines) > 1
        and old_lines[1].rstrip() != new_lines[1].rstrip()
    )
    if header_changed or subheader_changed:
        entries.append({
            "folder": "-", "file": fname,
            "diff_type": "列ヘッダの変化",
            "old_val": "chksimymlに記載のNode名称",
            "new_val": "SWC名/Node名に統一",
            "cause": "chksimyml のフォーマット変更(2ｰ①)に伴う変化点",
        })
    return entries


# ──────────────────────────────────────────
# 全差分抽出メイン
# ──────────────────────────────────────────

def extract_all(old_dir: Path, new_dir: Path) -> tuple[list[dict], dict]:
    """全ファイルの差分を抽出してエントリリストと検出統計を返す。

    Returns:
        entries: 抽出されたエントリリスト
        stats:   {
            "total_differing": int,   # 差分ありファイル数
            "detected":        int,   # 差分ありかつエントリ生成済みファイル数
            "undetected":      list,  # 差分ありなのに0件だったファイルラベル
        }
    """
    entries: list[dict] = []
    undetected: list[str] = []
    total_differing = 0
    detected_count  = 0

    def _file_differs(rel: str) -> bool:
        op, np = old_dir / rel, new_dir / rel
        return op.exists() and np.exists() and op.read_bytes() != np.read_bytes()

    def _check(label: str, new_e: list[dict], differs: bool) -> None:
        """差分あり/なしを記録し、差分ありなのに0件なら undetected に追加する"""
        nonlocal total_differing, detected_count
        if differs:
            total_differing += 1
            if new_e:
                detected_count += 1
            else:
                undetected.append(label)
        entries.extend(new_e)

    # [1] input_info.txt
    _check("input_info.txt",
           extract_input_info(old_dir, new_dir),
           _file_differs("input_info.txt"))

    # [2] グラフ画像（一括呼び出し後、gdir/sub 単位でチェック）
    img_entries = extract_image_diffs(old_dir, new_dir)
    entries += img_entries
    _GRAPH_DIRS = [
        "processing_load_duration_graph",
        "Sequence_duration_graph",
        "SWC_budget_duration_graph",
        "WakeupInterval_graph",
    ]
    for gdir in _GRAPH_DIRS:
        for sub in ("PASS", "FAIL"):
            old_imgs = image_list(old_dir / gdir / sub)
            new_imgs = image_list(new_dir / gdir / sub)
            if old_imgs != new_imgs:
                total_differing += 1
                if any(x["folder"] == gdir for x in img_entries):
                    detected_count += 1
                else:
                    undetected.append(f"{gdir}/{sub}")

    # [3] cpuload_requirements
    for fn in ("before_cpuload_requirements.csv", "after_cpuload_requirements.csv"):
        _check(f"temp/{fn}",
               extract_cpuload_requirements(old_dir, new_dir, fn),
               _file_differs(f"temp/{fn}"))

    # [4] requirements
    for fn in ("before_requirements.csv", "after_requirements.csv"):
        _check(f"temp/{fn}",
               extract_requirements_csv(old_dir, new_dir, fn),
               _file_differs(f"temp/{fn}"))

    # [5] budget
    _check("temp/before_budget.csv",
           extract_budget_csv(old_dir, new_dir, "before_budget.csv"),
           _file_differs("temp/before_budget.csv"))

    # [6] input_data
    for fn in ("input_data_ba.csv", "input_data_igr.csv"):
        _check(fn,
               extract_input_data_csv(old_dir, new_dir, fn),
               _file_differs(fn))

    # [7] tsync
    for fn in ("before_csv_data_tsync_PlusBA.csv", "after_csv_data_tsync_PlusBA.csv"):
        _check(f"temp/{fn}",
               extract_tsync_csv(old_dir, new_dir, fn),
               _file_differs(f"temp/{fn}"))

    # [8] processing_time_result
    _check("processing_time_result.csv",
           extract_processing_time_result(old_dir, new_dir),
           _file_differs("processing_time_result.csv"))

    # [9] schedule_result
    _check("schedule_result.csv",
           extract_schedule_result(old_dir, new_dir),
           _file_differs("schedule_result.csv"))

    # [10] schedule_fail_list
    _check("schedule_result_fail_list.csv",
           extract_schedule_fail_list(old_dir, new_dir),
           _file_differs("schedule_result_fail_list.csv"))

    return entries, {
        "total_differing": total_differing,
        "detected":        detected_count,
        "undetected":      undetected,
    }


# ──────────────────────────────────────────
# CSV 書き込み
# ──────────────────────────────────────────

CSV_COLUMNS = ["No", "フォルダ", "ファイル", "差分概要", "旧値", "新値", "リンク", "project", "variant", "推定原因"]


def write_to_csv(csv_path: Path, project: str, variant: str, entries: list[dict]) -> int:
    """差分エントリを CSV ファイルに書き出す（utf-8-sig: Excel で文字化けしない BOM 付き）"""
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for i, e in enumerate(entries, start=1):
            writer.writerow({
                "No": i,
                "フォルダ": e["folder"],
                "ファイル": e["file"],
                "差分概要": e["diff_type"],
                "旧値": e["old_val"],
                "新値": e["new_val"],
                "リンク": LINK_TEXT,
                "project": project,
                "variant": variant,
                "推定原因": e.get("cause") or "",
            })
    return len(entries)


# ──────────────────────────────────────────
# Excel 書き込み
# ──────────────────────────────────────────

def get_last_data_row(ws) -> int:
    # A列は連番数式（=A{n}+1）が残存することがあるため B列以降にデータがある行のみを対象とする
    last = 3
    for row in ws.iter_rows(min_row=4):
        if any(cell.value for cell in row[1:]):
            last = row[0].row
    return last


def _delete_ghost_rows(ws) -> None:
    """A列に数式が残るがB列以降が全て空の行（ゴースト行）を後ろから削除する"""
    ghost = [
        row[0].row for row in ws.iter_rows(min_row=4)
        if row[0].value is not None and not any(cell.value for cell in row[1:])
    ]
    for r in reversed(ghost):
        ws.delete_rows(r)


def write_to_excel(xlsx_path: Path, project: str, variant: str,
                   entries: list[dict], dry_run: bool) -> int:
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]

    # 既存エントリ（このバリアント）を削除して書き直す場合の確認
    existing = [
        row[0].row for row in ws.iter_rows(min_row=4)
        if row[7].value == project and row[8].value == variant
    ]
    if existing and not dry_run:
        print(f"  既存エントリ {len(existing)} 件を上書きします（行 {existing[0]}〜{existing[-1]}）")
        for r in reversed(existing):
            ws.delete_rows(r)

    # 削除後にA列数式のみ残ったゴースト行を除去する
    if not dry_run:
        _delete_ghost_rows(ws)

    last_row = get_last_data_row(ws)
    start_row = last_row + 1

    if dry_run:
        print(f"\n[DRY-RUN] 以下の {len(entries)} 件を行 {start_row}〜 に書き込む予定:")
        for i, e in enumerate(entries):
            print(f"  [{i+1:3}] folder={e['folder']} file={e['file']} "
                  f"diff={e['diff_type']} old={e['old_val']!r} new={e['new_val']!r}")
        return len(entries)

    for i, e in enumerate(entries):
        r = start_row + i
        ws.cell(row=r, column=1).value = f"=A{r-1}+1" if r > start_row else f"=A{last_row}+1"
        ws.cell(row=r, column=2).value = e["folder"]
        ws.cell(row=r, column=3).value = e["file"]
        ws.cell(row=r, column=4).value = e["diff_type"]
        ws.cell(row=r, column=5).value = e["old_val"]
        ws.cell(row=r, column=6).value = e["new_val"]
        ws.cell(row=r, column=7).value = LINK_TEXT
        ws.cell(row=r, column=8).value = project
        ws.cell(row=r, column=9).value = variant
        ws.cell(row=r, column=10).value = e.get("cause")

    wb.save(xlsx_path)
    return len(entries)


# ──────────────────────────────────────────
# CLI エントリポイント
# ──────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="差分抽出・Excel転記スクリプト")
    parser.add_argument("--project", required=True, help="プロジェクト名 (例: ver1)")
    parser.add_argument("--variant", required=True, help="バリアント名 (例: m02)")
    parser.add_argument("--old", required=True, help="旧解析結果ディレクトリ")
    parser.add_argument("--new", required=True, help="新解析結果ディレクトリ")
    parser.add_argument("--xlsx", required=True, help="Excelファイルパス")
    parser.add_argument("--dry-run", action="store_true", help="書き込まずに抽出結果のみ表示")
    parser.add_argument(
        "--csv-out",
        nargs="?",
        const="diff_output.csv",
        default=None,
        metavar="FILE",
        help="差分エントリを CSV ファイルに出力する。ファイル名省略時は diff_output.csv に出力（Excel 書き込みは行わない）",
    )
    args = parser.parse_args()

    old_dir = Path(args.old)
    new_dir = Path(args.new)
    xlsx_path = Path(args.xlsx)

    # --dry-run と --csv-out の同時指定チェック
    if args.dry_run and args.csv_out is not None:
        print("警告: --dry-run と --csv-out が同時指定されました。--csv-out を優先します。")
        args.dry_run = False

    for p in (old_dir, new_dir, xlsx_path):
        if not p.exists():
            print(f"エラー: パスが存在しません: {p}")
            raise SystemExit(1)

    print(f"差分抽出: {args.project}/{args.variant}")
    print(f"  旧: {old_dir}")
    print(f"  新: {new_dir}")

    entries, stats = extract_all(old_dir, new_dir)
    total_diff  = stats["total_differing"]
    detected    = stats["detected"]
    undetected  = stats["undetected"]
    rate        = (detected / total_diff * 100) if total_diff > 0 else 100.0

    print(f"  抽出件数: {len(entries)} 件")
    print(f"  検出率:   {detected}/{total_diff} ({rate:.0f}%)")

    if undetected:
        print(f"\n[FATAL] 差分ありなのに抽出ロジックが 0 件を返したファイルが {len(undetected)} 件あります:")
        for f in undetected:
            print(f"  - {f}")
        print("  → 新しい差分パターンが含まれている可能性があります。")
        print("  → scripts/extract_and_write_diff.py の抽出ロジックを更新してから再実行してください。")
        raise SystemExit(2)

    if args.csv_out is not None:
        csv_path = Path(args.csv_out)
        written = write_to_csv(csv_path, args.project, args.variant, entries)
        print(f"  CSV出力完了: {written} 件 → {csv_path}")
    else:
        written = write_to_excel(xlsx_path, args.project, args.variant, entries, args.dry_run)
        if not args.dry_run:
            print(f"  Excel転記完了: {written} 件")


if __name__ == "__main__":
    main()
