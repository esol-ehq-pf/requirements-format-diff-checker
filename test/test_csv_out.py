"""
tests/test_csv_out.py
Issue #1: --csv-out オプションのテストケース

TC-CSV-1: --csv-out FILE 指定時に指定パスに CSV が出力され Excel は変更されない
TC-CSV-2: --csv-out（ファイル名省略）時に diff_output.csv が出力される
TC-CSV-3: CSV の列ヘッダが SPEC の定義通りか
TC-CSV-4: CSV の各行の値が entries に対応しているか（No 連番・空 cause の扱いを含む）
TC-CSV-5: --dry-run + --csv-out 同時指定時に --csv-out が優先され CSV が出力される
TC-CSV-6: エントリ 0 件時でも CSV が出力される（ヘッダのみ）
"""
import csv
import sys
from pathlib import Path

import pytest

# スクリプトを直接 import できるようにパスを追加
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
from extract_and_write_diff import write_to_csv, CSV_COLUMNS, LINK_TEXT  # noqa: E402


# ─────────────────────────────────────────
# フィクスチャ
# ─────────────────────────────────────────

SAMPLE_ENTRIES = [
    {
        "folder": "-",
        "file": "input_info.txt",
        "diff_type": "入力ファイルの差分",
        "old_val": "-",
        "new_val": "要件チェックツール差分",
        "cause": "ver3ツールを用いたことによる差分",
    },
    {
        "folder": "tmp",
        "file": "before_requirements",
        "diff_type": "キーの追加",
        "old_val": "-",
        "new_val": "Targetキーが追加",
        "cause": None,  # None → 空文字に変換されること
    },
]


# ─────────────────────────────────────────
# TC-CSV-1: 指定ファイルに出力・件数を返す
# ─────────────────────────────────────────

def test_csv_out_named_file(tmp_path):
    """--csv-out result.csv 相当: 指定パスに CSV が作成され件数が返り、Excel は変更されない"""
    import openpyxl
    out = tmp_path / "result.csv"

    # Excel ファイルを用意し、書き込み前のサイズを記録
    xlsx = tmp_path / "dummy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output差分"
    wb.save(xlsx)
    size_before = xlsx.stat().st_size

    count = write_to_csv(out, "ver1", "m02", SAMPLE_ENTRIES)

    assert out.exists(), "指定パスに CSV が作成されること"
    assert count == len(SAMPLE_ENTRIES), "戻り値がエントリ件数と一致すること"
    # write_to_csv は Excel を触らないので xlsx のサイズは変わらない
    assert xlsx.stat().st_size == size_before, "write_to_csv は Excel を変更しないこと"


# ─────────────────────────────────────────
# TC-CSV-2: ファイル名省略時のデフォルト名はスクリプト側（main）で決まるので
#           write_to_csv 自体は任意パスを受け取れることを確認
# ─────────────────────────────────────────

def test_csv_out_default_filename(tmp_path, monkeypatch, capsys):
    """--csv-out のみ（ファイル名省略）時に const 値 diff_output.csv が使われること"""
    import openpyxl
    from extract_and_write_diff import main

    old_dir = tmp_path / "old"
    new_dir = tmp_path / "new"
    old_dir.mkdir()
    new_dir.mkdir()
    xlsx = tmp_path / "dummy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output差分"
    wb.save(xlsx)

    # --csv-out のみ（値なし）で呼び出す
    monkeypatch.setattr(
        sys, "argv",
        [
            "extract_and_write_diff.py",
            "--project", "ver1",
            "--variant", "m02",
            "--old", str(old_dir),
            "--new", str(new_dir),
            "--xlsx", str(xlsx),
            "--csv-out",          # 値なし → const="diff_output.csv" が使われる
        ],
    )
    # カレントディレクトリを tmp_path に変更して diff_output.csv の出力先を制御
    monkeypatch.chdir(tmp_path)
    main()

    default_csv = tmp_path / "diff_output.csv"
    assert default_csv.exists(), "ファイル名省略時は diff_output.csv に出力されること"


# ─────────────────────────────────────────
# TC-CSV-3: 列ヘッダ
# ─────────────────────────────────────────

def test_csv_columns_header(tmp_path):
    """CSV のヘッダが SPEC 定義の CSV_COLUMNS と完全一致すること"""
    out = tmp_path / "out.csv"
    write_to_csv(out, "ver1", "m02", SAMPLE_ENTRIES)

    with open(out, encoding="utf-8-sig", newline="") as f:
        header = next(csv.reader(f))

    assert header == CSV_COLUMNS, f"期待: {CSV_COLUMNS}, 実際: {header}"


# ─────────────────────────────────────────
# TC-CSV-4: 値の対応・No 連番・cause None → 空文字
# ─────────────────────────────────────────

def test_csv_row_values(tmp_path):
    """各行の値が entries の内容・連番・LINK_TEXT・project/variant と対応すること"""
    out = tmp_path / "out.csv"
    write_to_csv(out, "ver1", "m02", SAMPLE_ENTRIES)

    with open(out, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == len(SAMPLE_ENTRIES)

    # 1行目
    r0 = rows[0]
    assert r0["No"] == "1", "No 連番は 1 始まり"
    assert r0["フォルダ"] == SAMPLE_ENTRIES[0]["folder"]
    assert r0["ファイル"] == SAMPLE_ENTRIES[0]["file"]
    assert r0["差分概要"] == SAMPLE_ENTRIES[0]["diff_type"]
    assert r0["旧値"] == SAMPLE_ENTRIES[0]["old_val"]
    assert r0["新値"] == SAMPLE_ENTRIES[0]["new_val"]
    assert r0["リンク"] == LINK_TEXT
    assert r0["project"] == "ver1"
    assert r0["variant"] == "m02"
    assert r0["推定原因"] == SAMPLE_ENTRIES[0]["cause"]

    # 2行目: cause が None → 空文字
    r1 = rows[1]
    assert r1["No"] == "2", "No 連番は 2"
    assert r1["推定原因"] == "", "cause=None は空文字に変換されること"


# ─────────────────────────────────────────
# TC-CSV-5: --dry-run + --csv-out 同時指定 → --csv-out 優先（CLI 統合テスト）
# ─────────────────────────────────────────

def test_dry_run_and_csv_out_conflict(tmp_path, monkeypatch, capsys):
    """--dry-run と --csv-out の同時指定時に --csv-out が優先され CSV が出力されること"""
    import openpyxl
    from extract_and_write_diff import main

    # ダミーのディレクトリと xlsx を用意
    old_dir = tmp_path / "old"
    new_dir = tmp_path / "new"
    old_dir.mkdir()
    new_dir.mkdir()
    xlsx = tmp_path / "dummy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output差分"
    wb.save(xlsx)

    csv_out = tmp_path / "conflict.csv"

    monkeypatch.setattr(
        sys, "argv",
        [
            "extract_and_write_diff.py",
            "--project", "ver1",
            "--variant", "m02",
            "--old", str(old_dir),
            "--new", str(new_dir),
            "--xlsx", str(xlsx),
            "--dry-run",
            "--csv-out", str(csv_out),
        ],
    )
    main()

    captured = capsys.readouterr()
    assert "警告" in captured.out, "--dry-run + --csv-out 同時指定の警告が出力されること"
    assert csv_out.exists(), "--csv-out が優先され CSV が作成されること"


# ─────────────────────────────────────────
# TC-CSV-6: 0 件時はヘッダのみ出力
# ─────────────────────────────────────────

def test_csv_empty_entries(tmp_path):
    """エントリ 0 件時でもヘッダのみ含む CSV が作成されること"""
    out = tmp_path / "empty.csv"
    count = write_to_csv(out, "ver1", "m02", [])

    assert out.exists()
    assert count == 0

    with open(out, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows == [], "データ行は 0 件"


# ─────────────────────────────────────────
# TC-CSV-7: 出力先ディレクトリ不在時は FileNotFoundError
# ─────────────────────────────────────────

def test_csv_directory_not_found(tmp_path):
    """出力先ディレクトリが存在しない場合は FileNotFoundError が発生すること"""
    nonexistent = tmp_path / "no_such_dir" / "out.csv"
    with pytest.raises(FileNotFoundError):
        write_to_csv(nonexistent, "ver1", "m02", SAMPLE_ENTRIES)


# ─────────────────────────────────────────
# TC-CSV-8: --csv-out 使用時でも FATAL(exit=2) が発生すること
# ─────────────────────────────────────────

def test_csv_out_fatal_on_undetected(tmp_path, monkeypatch):
    """差分ありなのにエントリ 0 件のファイルがある場合、--csv-out 時でも exit=2(FATAL) で終了すること"""
    import openpyxl
    from extract_and_write_diff import main

    old_dir = tmp_path / "old"
    new_dir = tmp_path / "new"
    old_dir.mkdir()
    new_dir.mkdir()

    # input_info.txt: 内容が異なる（差分あり）→ 0 件になる状況を作る
    # ここでは extract_all をモックして undetected を強制的に発生させる
    import extract_and_write_diff as mod
    original_extract_all = mod.extract_all

    def mock_extract_all(old, new):
        entries, stats = original_extract_all(old, new)
        # 差分ありなのに 0 件のファイルがあったと見なす
        stats["total_differing"] = 1
        stats["detected"] = 0
        stats["undetected"] = ["fake_file.csv"]
        return entries, stats

    monkeypatch.setattr(mod, "extract_all", mock_extract_all)

    xlsx = tmp_path / "dummy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output差分"
    wb.save(xlsx)

    csv_out = tmp_path / "out.csv"

    monkeypatch.setattr(
        sys, "argv",
        [
            "extract_and_write_diff.py",
            "--project", "ver1",
            "--variant", "m02",
            "--old", str(old_dir),
            "--new", str(new_dir),
            "--xlsx", str(xlsx),
            "--csv-out", str(csv_out),
        ],
    )

    with pytest.raises(SystemExit) as exc_info:
        main()
    assert exc_info.value.code == 2, "--csv-out 時でも FATAL は exit=2 であること"
    assert not csv_out.exists(), "FATAL 終了時は CSV が出力されないこと"


# ─────────────────────────────────────────
# TC-CSV-9: --csv-out 時に --xlsx が存在しない場合は exit=1 で終了すること (NG-1)
# ─────────────────────────────────────────

def test_csv_out_xlsx_not_found_exits_1(tmp_path, monkeypatch):
    """--csv-out 指定時でも --xlsx が存在しない場合は exit=1 で終了すること"""
    import sys
    from extract_and_write_diff import main

    old_dir = tmp_path / "old"
    new_dir = tmp_path / "new"
    old_dir.mkdir()
    new_dir.mkdir()
    # xlsx は作成しない
    xlsx = tmp_path / "nonexistent.xlsx"
    csv_out = tmp_path / "out.csv"

    monkeypatch.setattr(
        sys, "argv",
        [
            "extract_and_write_diff.py",
            "--project", "ver1",
            "--variant", "m02",
            "--old", str(old_dir),
            "--new", str(new_dir),
            "--xlsx", str(xlsx),
            "--csv-out", str(csv_out),
        ],
    )

    with pytest.raises(SystemExit) as exc_info:
        main()
    assert exc_info.value.code == 1, "--xlsx が存在しないときは exit=1 であること"
    assert not csv_out.exists(), "exit=1 終了時は CSV が出力されないこと"


# ─────────────────────────────────────────
# TC-CSV-10: CSV 出力完了メッセージが stdout に出力されること (NG-3)
# ─────────────────────────────────────────

def test_csv_out_stdout_message(tmp_path, monkeypatch, capsys):
    """CSV 出力完了時に 'CSV出力完了: N 件 → <パス>' が標準出力に表示されること"""
    import openpyxl
    from extract_and_write_diff import main

    old_dir = tmp_path / "old"
    new_dir = tmp_path / "new"
    old_dir.mkdir()
    new_dir.mkdir()

    xlsx = tmp_path / "dummy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output差分"
    wb.save(xlsx)

    csv_out = tmp_path / "out.csv"

    monkeypatch.setattr(
        sys, "argv",
        [
            "extract_and_write_diff.py",
            "--project", "ver1",
            "--variant", "m02",
            "--old", str(old_dir),
            "--new", str(new_dir),
            "--xlsx", str(xlsx),
            "--csv-out", str(csv_out),
        ],
    )

    main()
    captured = capsys.readouterr()
    assert "CSV出力完了" in captured.out, "stdout に 'CSV出力完了' が含まれること"
    assert str(csv_out) in captured.out, "stdout に出力先パスが含まれること"


# ─────────────────────────────────────────
# TC-CSV-11: Excel 書き込み成功時に 'Excel転記完了' が stdout に出力されること (NG-3)
# ─────────────────────────────────────────

def test_excel_write_stdout_message(tmp_path, monkeypatch, capsys):
    """Excel 書き込み完了時に 'Excel転記完了: N 件' が標準出力に表示されること"""
    import openpyxl
    from extract_and_write_diff import main

    old_dir = tmp_path / "old"
    new_dir = tmp_path / "new"
    old_dir.mkdir()
    new_dir.mkdir()

    xlsx = tmp_path / "dummy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output差分"
    wb.save(xlsx)

    monkeypatch.setattr(
        sys, "argv",
        [
            "extract_and_write_diff.py",
            "--project", "ver1",
            "--variant", "m02",
            "--old", str(old_dir),
            "--new", str(new_dir),
            "--xlsx", str(xlsx),
        ],
    )

    main()
    captured = capsys.readouterr()
    assert "Excel転記完了" in captured.out, "stdout に 'Excel転記完了' が含まれること"


# ─────────────────────────────────────────
# TC-CSV-12: 既存エントリ上書き時に '既存エントリ N 件を上書きします' が stdout に出力されること (NG-1)
# ─────────────────────────────────────────

def test_excel_overwrite_stdout_message(tmp_path, monkeypatch, capsys):
    """既存の同 project/variant エントリが存在するとき '既存エントリ N 件を上書きします' が表示されること"""
    import openpyxl
    from extract_and_write_diff import main, SHEET_NAME

    old_dir = tmp_path / "old"
    new_dir = tmp_path / "new"
    old_dir.mkdir()
    new_dir.mkdir()

    # 既存エントリ（project=ver1, variant=m02）を 1 件持つ xlsx を準備
    xlsx = tmp_path / "dummy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # ヘッダ相当（行1〜3はヘッダ想定）、行4に既存エントリを置く
    ws["A3"] = 0  # No 数式の基準行
    ws["A4"] = "=A3+1"
    ws["B4"] = "-"
    ws["C4"] = "existing_file.txt"
    ws["D4"] = "入力ファイルの差分"
    ws["E4"] = "old"
    ws["F4"] = "new"
    ws["G4"] = "link"
    ws["H4"] = "ver1"   # project
    ws["I4"] = "m02"    # variant
    ws["J4"] = ""
    wb.save(xlsx)

    monkeypatch.setattr(
        sys, "argv",
        [
            "extract_and_write_diff.py",
            "--project", "ver1",
            "--variant", "m02",
            "--old", str(old_dir),
            "--new", str(new_dir),
            "--xlsx", str(xlsx),
        ],
    )

    main()
    captured = capsys.readouterr()
    assert "上書き" in captured.out, "stdout に '上書き' が含まれること（既存エントリ上書きメッセージ）"
